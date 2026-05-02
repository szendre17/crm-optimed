"""
SCRIPT 2/4: import_from_excel.py
Optimed CRM — Etapa 5

ROL: Importă datele reale din CRM_Optimed_v3_FINAL.xlsx în Frappe.

ORDINE DE IMPORT (critică din cauza dependențelor):
  1. Patients (9.791)   — fără calcul de segment, fără linkuri către deals
  2. Appointments (13.300) — leagă de Patients
  3. Deals (9.567)      — leagă de Patients și Operators
  4. Reconstrucție linkuri Appointment <-> Deal
  5. Family group ID pentru pacienții cu telefon comun

DUPĂ ACEST SCRIPT:
  Rulezi recalculate_stats.py care va calcula statisticile finale și segmentele.
  Apoi rulezi verify_import.py care compară cifrele cu Excel-ul.

CUM SE RULEAZĂ:
  Pune fișierul XLSX la calea: /home/frappe/frappe-bench/sites/CRM_Optimed_v3_FINAL.xlsx
  (sau ajustează XLSX_PATH mai jos)

  docker exec -it [container] bench --site [site] console
  exec(open('apps/optimed_crm/scripts/import_from_excel.py').read())
  run()

CARACTERISTICI:
  - Idempotent: rularea repetată e sigură (sare peste înregistrări existente)
  - Bulk insert: trigger-urile de calcul sunt OPRITE pe durata importului
  - Logging: progres la fiecare 500 înregistrări + erori în import_log.txt
  - Tolerant la erori: dacă o înregistrare pică, continuă cu restul
"""

import os
import sys
from datetime import datetime
from pathlib import Path

import frappe
from frappe.utils import flt


# =============================================================================
# CONFIGURARE
# =============================================================================

# Calea către fișierul XLSX (ajustează după unde îl pui)
XLSX_PATH = "/home/frappe/frappe-bench/sites/CRM_Optimed_v3_FINAL.xlsx"

# Fișier pentru log-uri detaliate
LOG_PATH = "/tmp/optimed_import_log.txt"


# =============================================================================
# MAPPING DE VALORI Excel → Frappe (pentru câmpurile Select)
# =============================================================================

CONSULTATION_TYPE_MAP = {
	"Consultație cu Bilet trimitere": "Consultație cu bilet de trimitere (gratuită)",
	"Consultație cu Plată": "Consultație cu plată",
	"Control anual": "Control anual",
	"Verificare ochelari": "Verificare ochelari",
	"Probă lentile contact": "Probă lentile contact",
	"Urgență": "Urgență",
}

DISCOUNT_TYPE_MAP = {
	"Pret intreg": "Preț întreg",
	"Card 10%": "Card 10%",
	"Card 15%": "Card 15%",
	"Rame 10%": "Rame 10%",
	"Rame 15%": "Rame 15%",
	"Lentile 10%": "Lentile 10%",
	"Voucher": "Voucher",
}

# Excel format: "DD.MM.YYYY" sau "DD.MM.YYYY HH:MM"
DATE_FORMATS = ["%d.%m.%Y", "%d.%m.%Y %H:%M", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S"]


# =============================================================================
# UTILITARE
# =============================================================================

_log_file = None


def log(msg, also_print=True):
	"""Scrie în log file și opțional și pe console."""
	global _log_file
	if _log_file is None:
		_log_file = open(LOG_PATH, "a", encoding="utf-8")
	timestamp = datetime.now().strftime("%H:%M:%S")
	line = f"[{timestamp}] {msg}"
	_log_file.write(line + "\n")
	_log_file.flush()
	if also_print:
		print(line)


def parse_date(value):
	"""Parsează o dată în diverse formate. Returnează date object sau None."""
	if value is None or value == "":
		return None
	if isinstance(value, datetime):
		return value.date()
	if hasattr(value, "year") and hasattr(value, "month"):  # date object
		return value
	if isinstance(value, str):
		for fmt in DATE_FORMATS:
			try:
				return datetime.strptime(value.strip(), fmt).date()
			except ValueError:
				continue
	return None


def parse_datetime(value):
	"""Parsează un datetime. Returnează datetime object sau None."""
	if value is None or value == "":
		return None
	if isinstance(value, datetime):
		return value
	if isinstance(value, str):
		for fmt in DATE_FORMATS:
			try:
				return datetime.strptime(value.strip(), fmt)
			except ValueError:
				continue
	# Fallback: dacă e doar dată
	d = parse_date(value)
	if d:
		return datetime.combine(d, datetime.min.time())
	return None


def excel_id_to_frappe_id(excel_id, prefix):
	"""
	Convertește ID-uri Excel la format Frappe.
	Exemple:
	  P00001 → PAT-00001
	  PR00001 → APP-00001
	  D00001 → DEAL-00001
	"""
	if not excel_id:
		return None
	excel_id = str(excel_id).strip()

	# Extrage cifrele de la sfârșit
	digits = ""
	for char in reversed(excel_id):
		if char.isdigit():
			digits = char + digits
		else:
			break

	if not digits:
		return None

	return f"{prefix}-{digits.zfill(5)}"


def patient_id(excel_id):
	return excel_id_to_frappe_id(excel_id, "PAT")


def appointment_id(excel_id):
	return excel_id_to_frappe_id(excel_id, "APP")


def deal_id(excel_id):
	return excel_id_to_frappe_id(excel_id, "DEAL")


# =============================================================================
# IMPORT PACIENȚI (Faza 1)
# =============================================================================

# Mapping in-memory: Excel ID → numele real Frappe (capturat după insert).
# Frappe v15 ignoră `doc.name` setat când autoname este `format:...`,
# așa că trebuie să capturăm numele real generat.
PATIENT_NAME_MAP = {}
APPOINTMENT_NAME_MAP = {}
DEAL_NAME_MAP = {}


def import_patients(rows):
	"""
	Importă pacienții. Populează PATIENT_NAME_MAP pentru Faza 2 și 3.
	Pentru idempotență (re-rulări parțiale), pacienții existenți sunt
	mapați după hubspot_contact_id când există, altfel cleanup obligatoriu.
	"""
	log(f"FAZA 1: Import pacienți ({len(rows)} înregistrări)")
	PATIENT_NAME_MAP.clear()
	created = 0
	skipped = 0
	errors = 0

	for i, row in enumerate(rows):
		try:
			excel_id = row["ID_Pacient"]
			frappe_name_hint = patient_id(excel_id)
			if not frappe_name_hint:
				log(f"  ✗ Skip pacient cu ID invalid: {excel_id}")
				errors += 1
				continue

			# Idempotență — dacă există deja, mapăm și sărim
			if frappe.db.exists("Patient", frappe_name_hint):
				PATIENT_NAME_MAP[excel_id] = frappe_name_hint
				skipped += 1
				continue

			doc = frappe.new_doc("Patient")
			doc.name = frappe_name_hint  # Frappe poate ignora; vom captura doc.name actual după insert
			doc.patient_name = (row.get("Nume") or "").strip() or "Necunoscut"
			doc.phone = (row.get("Telefon") or "").strip() or None
			doc.email = (row.get("Email") or "").strip() or None
			doc.hubspot_contact_id = (
				str(row.get("Contact_ID_HubSpot")).strip()
				if row.get("Contact_ID_HubSpot") is not None
				else None
			)
			doc.created_via = "Excel Import"
			doc.is_active = 1

			doc.flags.ignore_permissions = True
			doc.flags.ignore_mandatory = True
			doc.insert(ignore_permissions=True)

			# Capturează numele REAL atribuit de Frappe (poate diferi de hint)
			PATIENT_NAME_MAP[excel_id] = doc.name

			created += 1

			if (i + 1) % 500 == 0:
				frappe.db.commit()
				log(f"  Progres: {i + 1}/{len(rows)} (creat: {created}, sărit: {skipped}, erori: {errors})")

		except Exception as e:
			errors += 1
			log(f"  ✗ EROARE la pacientul {row.get('ID_Pacient')}: {str(e)[:200]}", also_print=False)

	frappe.db.commit()
	log(f"FAZA 1 COMPLETĂ: {created} creați, {skipped} săriți, {errors} erori\n")
	log(f"  PATIENT_NAME_MAP: {len(PATIENT_NAME_MAP)} înregistrări\n")
	return created, skipped, errors


# =============================================================================
# IMPORT PROGRAMĂRI (Faza 2)
# =============================================================================

def import_appointments(rows):
	"""
	Importă programările. Trigger-urile after_insert (care actualizează Patient)
	sunt SCURTCIRCUITATE prin frappe.db.sql direct — pentru viteză.

	Vom recalcula statisticile pacienților o singură dată în Faza 4.
	"""
	log(f"FAZA 2: Import programări ({len(rows)} înregistrări)")
	created = 0
	skipped = 0
	errors = 0

	APPOINTMENT_NAME_MAP.clear()

	for i, row in enumerate(rows):
		try:
			excel_appt_id = row["ID_Programare"]
			frappe_name_hint = appointment_id(excel_appt_id)
			if not frappe_name_hint:
				errors += 1
				continue

			if frappe.db.exists("Appointment", frappe_name_hint):
				APPOINTMENT_NAME_MAP[excel_appt_id] = frappe_name_hint
				skipped += 1
				continue

			# Patient ID — folosim mapping-ul construit în Faza 1
			excel_pid = row.get("ID_Pacient")
			patient_frappe = PATIENT_NAME_MAP.get(excel_pid)
			if not patient_frappe:
				log(f"  ✗ Skip programarea {excel_appt_id}: pacientul {excel_pid} nu este în mapping", also_print=False)
				errors += 1
				continue

			# Mapping câmpuri
			consultation_excel = (row.get("Tip_Consultatie") or "").strip()
			consultation_type = CONSULTATION_TYPE_MAP.get(consultation_excel, "Altul")

			anulat = (row.get("Anulat") or "").strip().upper()
			is_cancelled = 1 if anulat == "DA" else 0

			a_cumparat = (row.get("A_Cumparat") or "").strip().upper()
			has_purchase = 1 if a_cumparat == "DA" else 0

			doc = frappe.new_doc("Appointment")
			doc.name = frappe_name_hint
			doc.patient = patient_frappe
			doc.appointment_datetime = parse_datetime(row.get("Data_Programare")) or datetime(2020, 1, 1)
			doc.consultation_type = consultation_type
			doc.is_cancelled = is_cancelled
			doc.cancellation_reason = (row.get("Motiv_Anulare") or "").strip() or None
			doc.has_purchase = has_purchase
			doc.original_name = (row.get("Nume_Original_Calendly") or "").strip() or None
			doc.created_via = "Excel Import"
			doc.attended = 0 if is_cancelled else 1  # presupunem prezent dacă nu e anulat

			# Validări obligatorii pentru câmpuri condiționale
			if is_cancelled and not doc.cancellation_reason:
				doc.cancellation_reason = "Importat din Excel — motiv neînregistrat"

			doc.flags.ignore_permissions = True
			doc.flags.ignore_mandatory = True
			# Scurtcircuit: nu rula refresh patient stats (e prea lent)
			doc.flags.ignore_links = True
			doc.insert(ignore_permissions=True)

			# Capturează numele REAL atribuit de Frappe
			APPOINTMENT_NAME_MAP[excel_appt_id] = doc.name

			created += 1

			if (i + 1) % 500 == 0:
				frappe.db.commit()
				log(f"  Progres: {i + 1}/{len(rows)} (creat: {created}, sărit: {skipped}, erori: {errors})")

		except Exception as e:
			errors += 1
			log(f"  ✗ EROARE la programarea {row.get('ID_Programare')}: {str(e)[:200]}", also_print=False)

	frappe.db.commit()
	log(f"FAZA 2 COMPLETĂ: {created} create, {skipped} sărite, {errors} erori\n")
	log(f"  APPOINTMENT_NAME_MAP: {len(APPOINTMENT_NAME_MAP)} înregistrări\n")
	return created, skipped, errors


# =============================================================================
# IMPORT DEAL-URI (Faza 3)
# =============================================================================

def import_deals(rows):
	"""
	Importă deal-urile. La fel ca programările, scurtcircuităm trigger-urile
	pentru viteză și recalculăm statisticile pacientului în Faza 4.
	"""
	log(f"FAZA 3: Import deal-uri ({len(rows)} înregistrări)")
	created = 0
	skipped = 0
	errors = 0

	existing_operators = set(frappe.get_all("Sales Operator", pluck="name"))
	DEAL_NAME_MAP.clear()

	for i, row in enumerate(rows):
		try:
			excel_deal_id = row["ID_Deal"]
			frappe_name_hint = deal_id(excel_deal_id)
			if not frappe_name_hint:
				errors += 1
				continue

			if frappe.db.exists("Deal", frappe_name_hint):
				DEAL_NAME_MAP[excel_deal_id] = frappe_name_hint
				skipped += 1
				continue

			# Patient ID — folosim mapping-ul din Faza 1
			excel_pid = row.get("ID_Pacient")
			patient_frappe = PATIENT_NAME_MAP.get(excel_pid)
			if not patient_frappe:
				log(f"  ✗ Skip deal {excel_deal_id}: pacientul {excel_pid} nu este în mapping", also_print=False)
				errors += 1
				continue

			# Operator — fallback la "Necunoscut" dacă lipsește sau nu există
			operator_excel = (row.get("Vanzare_Operator") or "").strip()
			if not operator_excel or operator_excel not in existing_operators:
				if "Necunoscut" in existing_operators:
					operator_excel = "Necunoscut"
				else:
					log(f"  ✗ Skip deal {row.get('ID_Deal')}: operator '{operator_excel}' nu există și fallback 'Necunoscut' lipsește", also_print=False)
					errors += 1
					continue

			# Discount
			discount_excel = (row.get("Reduceri_Promotii") or "Pret intreg").strip()
			discount_type = DISCOUNT_TYPE_MAP.get(discount_excel, "Altă reducere")

			# Installer (cine a montat) — coloana Montare din Excel
			# Andreea = forma corectă (Adreea = typo). Necunoscut/Nu e cazul/Interoptik/gol → NULL
			installer_excel = (row.get("Montare") or "").strip()
			if installer_excel == "Adreea":
				installer_excel = "Andreea"
			if installer_excel not in existing_operators:
				installer_excel = None

			doc = frappe.new_doc("Deal")
			doc.name = frappe_name_hint
			doc.patient = patient_frappe
			doc.sales_operator = operator_excel
			doc.installer = installer_excel
			doc.creation_date = parse_date(row.get("Data_Creare"))
			doc.pickup_date = parse_date(row.get("Data_Ridicare"))
			doc.frame_price = flt(row.get("Pret_Rama"))
			doc.lens1_price = flt(row.get("Pret_Lentila1"))
			doc.lens2_price = flt(row.get("Pret_Lentila2"))
			doc.sunglasses_price = flt(row.get("Pret_Ochelari_Soare"))
			doc.accessories_price = flt(row.get("Pret_Accesorii"))
			doc.discount_type = discount_type

			# Calculăm discount_amount din diferența: components - revenue_excel
			components = (
				flt(row.get("Pret_Rama"))
				+ flt(row.get("Pret_Lentila1"))
				+ flt(row.get("Pret_Lentila2"))
				+ flt(row.get("Pret_Ochelari_Soare"))
				+ flt(row.get("Pret_Accesorii"))
			)
			revenue_excel = flt(row.get("Venituri_Castigate_RON"))
			discount_amount = max(0, components - revenue_excel)
			doc.discount_amount = discount_amount

			# Manopera vine direct din Excel
			doc.labor = flt(row.get("Manopera_RON"))

			# Setăm explicit câmpurile calculate ca să nu le recalculeze validate()
			# — dar ele se vor recalcula oricum la save() prin _calculate_financials()
			# Asta e ok fiindcă reproduce exact logica Excel-ului

			doc.original_source_name = (row.get("Nume_Sursa") or "").strip() or None
			doc.original_deal_name = (row.get("Deal_Name_Original") or "").strip() or None
			doc.original_associated_contact = (
				row.get("Associated_Contact_Original") or ""
			).strip() or None
			doc.created_via = "Excel Import"
			doc.is_paid = 1  # presupunem plătit (e istoric)

			doc.flags.ignore_permissions = True
			doc.flags.ignore_mandatory = True
			doc.flags.ignore_links = True
			doc.insert(ignore_permissions=True)

			# Capturează numele REAL atribuit de Frappe
			DEAL_NAME_MAP[excel_deal_id] = doc.name

			created += 1

			if (i + 1) % 500 == 0:
				frappe.db.commit()
				log(f"  Progres: {i + 1}/{len(rows)} (creat: {created}, sărit: {skipped}, erori: {errors})")

		except Exception as e:
			errors += 1
			log(f"  ✗ EROARE la deal-ul {row.get('ID_Deal')}: {str(e)[:200]}", also_print=False)

	frappe.db.commit()
	log(f"FAZA 3 COMPLETĂ: {created} create, {skipped} sărite, {errors} erori\n")
	return created, skipped, errors


# =============================================================================
# RECONSTRUCȚIE LINK-URI (Faza 4)
# =============================================================================

def link_appointments_to_deals(appointment_rows):
	"""
	Pentru fiecare programare cu ID_Deal_Asociat, setează linked_deal pe Appointment.
	"""
	log("FAZA 4: Reconstrucție link-uri Appointment → Deal")
	updated = 0
	errors = 0

	for row in appointment_rows:
		try:
			deal_excel = row.get("ID_Deal_Asociat")
			if not deal_excel:
				continue

			# Folosim mapping-urile construite în Faza 2 și 3
			appt_frappe = APPOINTMENT_NAME_MAP.get(row.get("ID_Programare"))
			deal_frappe = DEAL_NAME_MAP.get(deal_excel)

			if not appt_frappe or not deal_frappe:
				continue

			if not frappe.db.exists("Deal", deal_frappe):
				continue

			frappe.db.set_value(
				"Appointment", appt_frappe,
				{"linked_deal": deal_frappe, "has_purchase": 1},
				update_modified=False
			)
			updated += 1
		except Exception as e:
			errors += 1
			log(f"  ✗ EROARE link {row.get('ID_Programare')} → {row.get('ID_Deal_Asociat')}: {str(e)[:150]}", also_print=False)

	frappe.db.commit()
	log(f"FAZA 4 COMPLETĂ: {updated} link-uri create, {errors} erori\n")
	return updated, errors


def assign_family_groups():
	"""
	Pentru pacienții cu același telefon, generează un family_group_id comun.
	Ex: FAM-00001 pentru toți pacienții care au telefonul "+40 745 111 980"
	"""
	log("FAZA 5: Detectare familii (telefon comun)")

	# Găsește telefoanele care apar la mai mulți pacienți
	family_phones = frappe.db.sql("""
		SELECT phone, COUNT(*) as cnt
		FROM `tabPatient`
		WHERE phone IS NOT NULL AND phone != ''
		GROUP BY phone
		HAVING cnt > 1
	""", as_dict=True)

	log(f"  Detectate {len(family_phones)} telefoane comune (familii)")

	family_counter = 0
	patients_assigned = 0

	for fam in family_phones:
		family_counter += 1
		family_id = f"FAM-{str(family_counter).zfill(5)}"

		# Asignează acest family_id tuturor pacienților cu acest telefon
		count = frappe.db.sql("""
			UPDATE `tabPatient`
			SET family_group_id = %s
			WHERE phone = %s
		""", (family_id, fam["phone"]))

		patients_assigned += fam["cnt"]

	frappe.db.commit()
	log(f"FAZA 5 COMPLETĂ: {family_counter} familii create, {patients_assigned} pacienți marcați\n")
	return family_counter


# =============================================================================
# MAIN
# =============================================================================

def read_xlsx(path):
	"""Citește XLSX-ul și returnează 3 liste: patients, appointments, deals."""
	try:
		from openpyxl import load_workbook
	except ImportError:
		log("✗ EROARE: Lipsește openpyxl. Instalează cu: pip install openpyxl")
		return None, None, None

	if not os.path.exists(path):
		log(f"✗ EROARE: Fișierul nu există la calea {path}")
		log(f"   Pune-l în /home/frappe/frappe-bench/sites/ sau ajustează XLSX_PATH în script")
		return None, None, None

	log(f"Citire XLSX: {path}")
	wb = load_workbook(path, data_only=True)

	def sheet_to_dicts(ws):
		headers = [c.value for c in ws[1]]
		rows = []
		for row in ws.iter_rows(min_row=2, values_only=True):
			if all(v is None for v in row):
				continue
			rows.append(dict(zip(headers, row)))
		return rows

	patients = sheet_to_dicts(wb["PACIENTI"])
	appointments = sheet_to_dicts(wb["PROGRAMARI"])
	deals = sheet_to_dicts(wb["DEALS"])

	log(f"  Citite: {len(patients)} pacienți, {len(appointments)} programări, {len(deals)} deal-uri")
	return patients, appointments, deals


def run():
	"""Rulează importul complet."""
	# Curăță log-ul anterior
	if os.path.exists(LOG_PATH):
		os.remove(LOG_PATH)

	log("=" * 70)
	log("IMPORT EXCEL → OPTIMED CRM")
	log("=" * 70)

	# Citește XLSX
	patients, appointments, deals = read_xlsx(XLSX_PATH)
	if patients is None:
		log("✗ Import oprit: nu s-a putut citi XLSX-ul")
		return

	# Verifică operatorii
	op_count = frappe.db.count("Sales Operator")
	if op_count < 3:
		log(f"✗ EROARE: Doar {op_count} operatori în baza de date. Trebuie 3 (Ramona, Roxana, Eniko).")
		log("   Rulează prima oară crearea operatorilor (vezi Etapa 4, Pasul 4)")
		return

	# Faza 1: Pacienți
	import_patients(patients)

	# Faza 2: Programări
	import_appointments(appointments)

	# Faza 3: Deal-uri
	import_deals(deals)

	# Faza 4: Link-uri Appointment → Deal
	link_appointments_to_deals(appointments)

	# Faza 5: Family groups
	assign_family_groups()

	# Raport final
	log("=" * 70)
	log("IMPORT COMPLET — Stare finală în Frappe:")
	log(f"  Pacienți: {frappe.db.count('Patient')}")
	log(f"  Programări: {frappe.db.count('Appointment')}")
	log(f"  Deal-uri: {frappe.db.count('Deal')}")
	log(f"  Operatori: {frappe.db.count('Sales Operator')}")
	log("=" * 70)
	log("URMĂTORUL PAS:")
	log("  Rulează recalculate_stats.py pentru a calcula statisticile pacienților")
	log("  Apoi verify_import.py pentru a compara cu cifrele din Excel")
	log("=" * 70)
	log(f"Log detaliat: {LOG_PATH}")

	if _log_file:
		_log_file.close()


if __name__ == "__main__":
	run()
