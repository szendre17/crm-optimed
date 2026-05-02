"""
Backfill câmpul Deal.installer din coloana Montare a Excel-ului.

Strategie de matching îmbunătățită — folosește o cheie compozită mai precisă:
  (patient_name_normalized, creation_date, sales_operator, frame_price, lens1_price, lens2_price)

Asta rezolvă cazurile unde un pacient cumpără mai multe perechi în aceeași zi
cu același vânzător (cheia de bază nu era unică).

Idempotent: șterge installer-ul existent înainte de re-populare.
"""

import openpyxl
import frappe
from collections import defaultdict


XLSX_PATH = "/home/frappe/frappe-bench/sites/CRM_Optimed_v3_FINAL.xlsx"


def normalize_installer(value, valid_operators):
	if not value:
		return None
	val = str(value).strip()
	if val == "Adreea":
		val = "Andreea"
	if val in valid_operators:
		return val
	return None


def parse_excel_date(value):
	if not value:
		return None
	if hasattr(value, "date"):
		return value.date()
	import datetime
	for fmt in ["%d.%m.%Y", "%Y-%m-%d", "%d.%m.%Y %H:%M"]:
		try:
			return datetime.datetime.strptime(str(value).strip(), fmt).date()
		except (ValueError, AttributeError):
			pass
	return None


def to_float(value):
	"""Convertesc la float pentru a permite comparație cu valorile din DB."""
	if value is None or value == "":
		return 0.0
	try:
		return float(value)
	except (ValueError, TypeError):
		return 0.0


def normalize_name(s):
	if not s:
		return ""
	return str(s).strip().upper()


def run():
	print("=" * 70)
	print("BACKFILL Deal.installer — matching avansat")
	print("=" * 70)

	valid_operators = set(frappe.get_all("Sales Operator", pluck="name"))
	print(f"Sales Operators valizi: {sorted(valid_operators)}")

	# Reset installer pentru toate deal-urile (idempotent)
	frappe.db.sql("UPDATE `tabDeal` SET installer = NULL")
	frappe.db.commit()
	print("Reset installer pentru toate deal-urile.\n")

	# Citesc Excel
	wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)
	ws = wb["DEALS"]
	header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
	idx = {h: i for i, h in enumerate(header)}

	# Construiesc index Excel: cheie compozită → list of installers
	excel_idx = defaultdict(list)
	excel_count = 0
	for row in ws.iter_rows(min_row=2, values_only=True):
		excel_count += 1
		sales_op = (row[idx["Vanzare_Operator"]] or "").strip()
		if sales_op == "Adreea":
			sales_op = "Andreea"

		key = (
			normalize_name(row[idx["Nume_Sursa"]]),
			parse_excel_date(row[idx["Data_Creare"]]),
			sales_op,
			to_float(row[idx["Pret_Rama"]]),
			to_float(row[idx["Pret_Lentila1"]]),
			to_float(row[idx["Pret_Lentila2"]]),
		)
		installer = normalize_installer(row[idx["Montare"]], valid_operators)
		excel_idx[key].append(installer)

	print(f"Înregistrări Excel: {excel_count}")
	print(f"Chei compozite distincte: {len(excel_idx)}")

	# Match Frappe Deals — folosesc original_source_name (= Nume_Sursa din DEALS)
	# pentru că pentru pacienții-familie, numele pe deal poate diferi de patient_name
	deals = frappe.db.sql("""
		SELECT name, creation_date, sales_operator,
		       frame_price, lens1_price, lens2_price,
		       original_source_name
		FROM `tabDeal`
	""", as_dict=True)
	print(f"Deal-uri Frappe: {len(deals)}\n")

	matched = 0
	updated = 0
	unmatched = 0
	for d in deals:
		key = (
			normalize_name(d["original_source_name"]),
			d["creation_date"],
			d["sales_operator"],
			to_float(d["frame_price"]),
			to_float(d["lens1_price"]),
			to_float(d["lens2_price"]),
		)
		if key in excel_idx and excel_idx[key]:
			installer = excel_idx[key].pop(0)
			matched += 1
			if installer:
				frappe.db.set_value("Deal", d["name"], "installer", installer, update_modified=False)
				updated += 1
		else:
			unmatched += 1

	frappe.db.commit()

	print(f"Match-uri: {matched}/{len(deals)} ({100*matched/len(deals):.1f}%)")
	print(f"Updated cu installer (non-NULL): {updated}")
	print(f"Unmatched: {unmatched}\n")

	# Statistici finale
	print("=" * 70)
	print("DISTRIBUȚIE INSTALLER (Frappe vs Excel așteptat)")
	print("=" * 70)
	expected = {"Eniko": 2858, "Ramona": 1894, "Attila": 1549, "Roxana": 1493, "Andreea": 600}
	dist = frappe.db.sql("""
		SELECT COALESCE(installer, '(NULL)') as installer, COUNT(*) as n
		FROM `tabDeal`
		GROUP BY installer
		ORDER BY n DESC
	""", as_dict=True)
	for row in dist:
		exp = expected.get(row["installer"])
		if exp:
			diff = row["n"] - exp
			marker = "✓" if abs(diff) <= 5 else f"({diff:+d})"
			print(f"  {row['installer']:15s}: {row['n']:5d}  (Excel: {exp:5d}) {marker}")
		else:
			# Pentru NULL: așteptat 1173 (gol/Nu e cazul/Interoptik)
			if row["installer"] == "(NULL)":
				print(f"  {row['installer']:15s}: {row['n']:5d}  (Excel ~1173 cu Montare necunoscut)")
			else:
				print(f"  {row['installer']:15s}: {row['n']:5d}")
	print("=" * 70)

	return updated
